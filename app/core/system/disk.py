import openpyxl
import yadisk
import pandas as pd
import requests

from io import BytesIO

class YandexDisk:
    def __init__(self, id, secret, token):
        self.id = id
        self.secret = secret
        self.token = token
        self.client = None

    async def init_client(self):
        if self.client is None:
            self.client = yadisk.AsyncClient(id=self.id, secret=self.secret, token=self.token)

    async def read_file(self, file_path):
        await self.init_client()
        # Получаем ссылку на скачивание файла
        download_link = await self.client.get_download_link(file_path)

        # Получаем содержимое файла по ссылке
        response = requests.get(download_link)

        # Читаем Excel-файл из содержимого
        wb = openpyxl.load_workbook(BytesIO(response.content))

        return wb.active

    async def get_sources(self, file_path):
        sheet = await self.read_file(file_path)

        data = []
        for row in sheet.rows:
            data.append([cell.value for cell in row])

        df = pd.DataFrame(data)

        # Преобразуем DataFrame в список списков
        result = df.values.tolist() # название + ссылка
        # Получаем список ссылок
        urls = [item[1] for item in result if item[1] is not None] # просто ссылки

        print(f'Источники: {urls}')

        return urls
    
    async def get_ban_categories(self, file_path):
        sheet = await self.read_file(file_path)

        # Создаем пустой список для слов
        words = []

        # Итерируем по строкам листа
        for row in range(1, sheet.max_row + 1):
            # Получаем значение ячейки в текущей строке
            cell_value = sheet.cell(row=row, column=1).value

            # Если ячейка не пуста, добавляем значение в список слов
            if cell_value is not None:
                words.append(str(cell_value).strip())

        return words
